# -*- coding: utf-8 -*-
"""
cargar_piezas.py
Lee un Excel con columnas (Pieza, Descripcion) y lo envía al servidor
de escaneo para guardarlo en Supabase.

Formato esperado del Excel:
    Columna A: Código de Pieza
    Columna B: Descripción (opcional)

El header se detecta automáticamente en las primeras 20 filas.
"""
import json
import os
import sys
import urllib.request
import urllib.error


def leer_piezas_excel(ruta_excel):
    """Lee el Excel y devuelve [{codigo_pieza, descripcion}]."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    hdr_row = None
    cols = {}
    for r in range(1, 21):
        encontradas = {}
        for c in range(1, (ws.max_column or 2) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                u = v.strip().upper()
                u = u.replace('Í', 'I').replace('Ó', 'O').replace('É', 'E')
                u = u.replace('Á', 'A').replace('Ú', 'U').replace('Ñ', 'N')
                if 'PIEZA' in u or 'CODIGO' in u or 'COD' in u:
                    encontradas['pieza'] = c
                elif 'DESC' in u:
                    encontradas['desc'] = c
        if 'pieza' in encontradas:
            hdr_row = r
            cols = encontradas
            break

    if not hdr_row or 'pieza' not in cols:
        raise ValueError(
            'No se encontró la columna de Pieza/Código en el Excel.\n'
            'Asegurate de que la Columna A tenga "Pieza" o "Código" como header.')

    filas = []
    for r in range(hdr_row + 1, (ws.max_row or hdr_row) + 1):
        pieza = ws.cell(r, cols['pieza']).value
        if pieza is None or str(pieza).strip() == '':
            continue
        desc = ws.cell(r, cols.get('desc', 2)).value or ''
        filas.append({
            'codigo_pieza': str(pieza).strip(),
            'descripcion': str(desc).strip(),
        })

    return filas


def enviar_al_servidor(piezas, host='127.0.0.1', port=8000):
    """Envía las piezas al servidor POST /cargar_piezas."""
    url = f'http://{host}:{port}/cargar_piezas'
    data = json.dumps({'piezas': piezas}).encode('utf-8')
    req = urllib.request.Request(url, data=data, method='POST',
                                headers={'Content-Type': 'application/json'})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode('utf-8'))


def main():
    print()
    print('=' * 55)
    print('  CARGAR PIEZAS DESDE EXCEL')
    print('=' * 55)
    print()

    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    ruta = filedialog.askopenfilename(
        title='Selecciona el Excel de piezas',
        filetypes=[('Archivos Excel', '*.xlsx'), ('Todos los archivos', '*.*')],
    )
    root.destroy()

    if not ruta:
        print('  No se seleccionó ningún archivo.')
        return

    print()
    print('  Leyendo Excel...')
    piezas = leer_piezas_excel(ruta)
    print(f'  Se encontraron {len(piezas)} piezas.')

    if not piezas:
        print('  No hay piezas para cargar.')
        return

    print()
    print('  Primeras 5 piezas:')
    for p in piezas[:5]:
        desc = f' - {p["descripcion"]}' if p['descripcion'] else ''
        print(f'    {p["codigo_pieza"]}{desc}')
    if len(piezas) > 5:
        print(f'    ... y {len(piezas) - 5} más')

    print()
    confirmar = input('  ¿Cargar al servidor? (s/n): ').strip().lower()
    if confirmar != 's':
        print('  Cancelado.')
        return

    try:
        print()
        print('  Enviando al servidor...')
        resultado = enviar_al_servidor(piezas)
        if resultado.get('ok'):
            print(f'  ¡Listo! {resultado["insertadas"]} piezas cargadas.')
        else:
            print(f'  Error: {resultado}')
    except urllib.error.URLError as e:
        print(f'  No se pudo conectar al servidor: {e}')
        print('  Asegurate de que el servidor esté corriendo (opción 11 del menú).')
    except Exception as e:
        print(f'  Error: {e}')

    print()


if __name__ == '__main__':
    main()
