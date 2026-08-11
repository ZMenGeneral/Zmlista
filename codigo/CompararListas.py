# -*- coding: utf-8 -*-
"""
CompararListas.py
OPCION 7 (temporal) - Comparar 2 Excel de listas de precios:
    Lee dos archivos Excel con el formato de lista (CODIGO, DESCRIPCION,
    MODELO, MARCA, PRECIO, CANT), los combina por CODIGO y para cada codigo
    se queda con la fila que tenga la mayor CANT (cantidad).

    Los codigos que solo existen en un archivo se conservan tal cual.
    Genera "Lista Comparada DD-MM-AAAA.xlsx" en la carpeta salidas con una
    columna ORIGEN que indica de donde salio cada producto (Archivo 1 / 2).

Uso:
    python CompararListas.py                                    (abre explorador)
    python CompararListas.py "C:/ruta/lista1.xlsx" "C:/ruta/lista2.xlsx"
"""
import os
import sys
from datetime import date

import consola
from ConvertirListaExcel import parse_number, read_excel_rows

CARPETA_CODIGO = os.path.dirname(os.path.abspath(__file__))
CARPETA_PROYECTO = os.path.dirname(CARPETA_CODIGO)
CARPETA_SALIDAS = os.path.join(CARPETA_PROYECTO, 'salidas')


def ask_excel_dialog(titulo):
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title=titulo,
        filetypes=[('Archivos Excel', '*.xlsx'), ('Todos los archivos', '*.*')],
        initialdir=CARPETA_SALIDAS if os.path.isdir(CARPETA_SALIDAS) else CARPETA_PROYECTO,
    )
    root.destroy()
    return path


def _cant_num(cant):
    if isinstance(cant, str):
        cant = parse_number(cant)
    return cant if cant is not None else 0


def combinar(lista1, lista2):
    """Combina por codigo quedandose con la mayor cantidad.
    Devuelve (filas, resumen) donde cada fila tiene 'origen' (Archivo 1/2)."""
    codigos1 = set(r['codigo'] for r in lista1)
    codigos2 = set(r['codigo'] for r in lista2)

    filas = {}
    for origen, lista in (('Archivo 1', lista1), ('Archivo 2', lista2)):
        for r in lista:
            cod = r['codigo']
            cant = _cant_num(r['cant'])
            actual = filas.get(cod)
            if actual is None or cant > actual['cant']:
                nuevo = dict(r)
                nuevo['origen'] = origen
                nuevo['cant'] = cant
                filas[cod] = nuevo

    resumen = {
        'solo1': len(codigos1 - codigos2),
        'solo2': len(codigos2 - codigos1),
        'ambos': len(codigos1 & codigos2),
    }
    return list(filas.values()), resumen


def guardar_excel(filas, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    encabezados = ['CODIGO', 'DESCRIPCION', 'MODELO', 'MARCA', 'PRECIO', 'CANT', 'ORIGEN']
    anchos = [18, 45, 20, 18, 14, 10, 12]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparada'
    for col, h in enumerate(encabezados, 1):
        cel = ws.cell(1, col, h)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor='1F4E78')
        cel.alignment = Alignment(horizontal='center')
    for r, fila in enumerate(filas, 2):
        ws.cell(r, 1, fila['codigo'])
        ws.cell(r, 2, fila.get('desc'))
        ws.cell(r, 3, fila.get('modelo'))
        ws.cell(r, 4, fila.get('marca'))
        ws.cell(r, 5, fila.get('precio'))
        ws.cell(r, 6, fila.get('cant'))
        ws.cell(r, 7, fila.get('origen'))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(encabezados))}{len(filas) + 1}'
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    wb.save(out_path)
    return out_path


def main():
    args = sys.argv[1:]

    consola.titulo('OPCION 7 (TEMPORAL) - COMPARAR 2 EXCEL POR CODIGO Y MAYOR CANT')

    if len(args) >= 2:
        p1, p2 = args[0], args[1]
    else:
        p1 = ask_excel_dialog('Selecciona el PRIMER Excel de lista')
        if not p1:
            print('  No seleccionaste ningun archivo. Cancelado.')
            return
        p2 = ask_excel_dialog('Selecciona el SEGUNDO Excel de lista')
        if not p2:
            print('  No seleccionaste ningun archivo. Cancelado.')
            return

    for p in (p1, p2):
        if not os.path.exists(p):
            print(f'  No se encontro el archivo: {p}')
            return

    print(f'  Archivo 1: {p1}')
    print(f'  Archivo 2: {p2}')

    try:
        lista1 = read_excel_rows(p1)
    except Exception as e:
        print(f'  Error al leer el Excel 1: {e}')
        return
    try:
        lista2 = read_excel_rows(p2)
    except Exception as e:
        print(f'  Error al leer el Excel 2: {e}')
        return

    if not lista1 and not lista2:
        print('  Los dos archivos estan vacios. Cancelado.')
        return

    filas, resumen = combinar(lista1, lista2)

    ganadas1 = sum(1 for f in filas if f['origen'] == 'Archivo 1')
    ganadas2 = sum(1 for f in filas if f['origen'] == 'Archivo 2')

    os.makedirs(CARPETA_SALIDAS, exist_ok=True)
    nombre = 'Lista Comparada ' + date.today().strftime('%d-%m-%Y') + '.xlsx'
    out = os.path.join(CARPETA_SALIDAS, nombre)
    guardar_excel(filas, out)

    print()
    print('-' * 60)
    print(f'  Productos del Archivo 1: {consola.negrita(len(lista1))}')
    print(f'  Productos del Archivo 2: {consola.negrita(len(lista2))}')
    print(f'  Solo en Archivo 1: {consola.cian(resumen["solo1"])}')
    print(f'  Solo en Archivo 2: {consola.cian(resumen["solo2"])}')
    print(f'  En ambos archivos: {consola.cian(resumen["ambos"])}')
    print(f'  Ganancias Archivo 1 (mayor CANT): {consola.verde(ganadas1)}')
    print(f'  Ganancias Archivo 2 (mayor CANT): {consola.verde(ganadas2)}')
    print('-' * 60)
    print(consola.verde(f'  OK: {len(filas)} productos comparados -> {out}'))
    consola.separador()


if __name__ == '__main__':
    main()
    try:
        input('Presiona Enter para cerrar...')
    except EOFError:
        pass
