# -*- coding: utf-8 -*-
"""
ConvertirListaExcel.py

OPCION 1 - Lista de precios (TXT -> Excel):
    Convierte listas de precios en TXT (ancho fijo, Windows-1252) y rellena la
    plantilla "LISTA base.xlsx" generando "Lista Zm DD-MM-AAAA.xlsx".

OPCION 2 - Lista en Bs (Excel -> Excel):
    Lee un Excel de lista de precios (normalmente el generado en la opcion 1),
    rellena la plantilla "LISTA Base Bs.xlsx" generando "Lista Zm Bs DD-MM-AAAA.xlsx".
    Pide el monto del dia (Z), multiplica la columna PRECIO por Z y reemplaza la
    X del mensaje "COMPRAS A PARTIR DE: X Bs" por 450 * Z.

Uso:
    python ConvertirListaExcel.py                          (opcion 1, abre explorador)
    python ConvertirListaExcel.py "C:/ruta/archivo.txt"
    python ConvertirListaExcel.py "C:/ruta/carpeta"        (convierte todos los *.txt)
    python ConvertirListaExcel.py --bs "C:/ruta/lista.xlsx" [Z]

Comportamiento opcion 1:
    - Quita productos con existencia 0
    - Limpia la categoria (solo para el orden; no se escribe en la plantilla)
    - Ordena por Categoria y luego por Codigo
    - Rellena la plantilla LISTA base:
        B = CODIGO, C = DESCRIPCION, D = MODELO (marca vehiculo),
        E = MARCA (marca proveedor), F = PRECIO, G = CANT (vacia)
    - Actualiza la fecha (celda B10) con la fecha de hoy

Comportamiento opcion 2:
    - Pide el archivo Excel de origen y el monto del dia (Z)
    - Rellena la plantilla LISTA Base Bs:
        B = CODIGO, C = DESCRIPCION, D = MODELO, E = MARCA,
        F = PRECIO x Z, G = CANT (se copia si existe)
    - Actualiza la fecha (B10) y el mensaje "COMPRAS A PARTIR DE: X Bs"
      con X = 450 x Z
"""
import copy
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date
from xml.sax.saxutils import escape

import consola

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_MC = 'http://schemas.openxmlformats.org/markup-compatibility/2006'
NS_X14AC = 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac'
NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
ET.register_namespace('', NS)
ET.register_namespace('mc', NS_MC)
ET.register_namespace('x14ac', NS_X14AC)
ET.register_namespace('r', NS_R)

TAIL_RE = re.compile(r'^\s*([\d.,]+)\s+([\d.,]+)\s*$')

CARPETA_CODIGO = os.path.dirname(os.path.abspath(__file__))
CARPETA_PROYECTO = os.path.dirname(CARPETA_CODIGO)
CARPETA_PLANTILLAS = os.path.join(CARPETA_PROYECTO, 'plantillas')
CARPETA_SALIDAS = os.path.join(CARPETA_PROYECTO, 'salidas')
LOG_FILAS = os.path.join(CARPETA_PROYECTO, 'Registro Filas de Items.txt')


def excel_a_pdf(excel_path):
    """Convierte un archivo Excel a PDF usando la COM de Excel."""
    try:
        import win32com.client
        import pythoncom
        pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
        pythoncom.CoInitialize()
        xl = win32com.client.DispatchEx('Excel.Application')
        try:
            wb = xl.Workbooks.Open(os.path.abspath(excel_path))
            wb.ExportAsFixedFormat(0, pdf_path)
            wb.Close(SaveChanges=False)
            return pdf_path
        finally:
            try:
                xl.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()
    except Exception as e:
        print(consola.amarillo(f'  (No se pudo generar el PDF: {e})'))
        return None


def parse_number(value):
    """Convierte '218,62' -> 218.62  y  '1.456,00' -> 1456.0. None si no es numero."""
    if not value or not value.strip():
        return None
    clean = value.strip().replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return None


def parse_monto(value):
    """Interpreta un monto ingresado por el usuario: '38.5', '38,5', '38,50',
    '1.456,50' o '1,456.50'."""
    if not value:
        return None
    s = value.strip().replace('Bs', '').replace('$', '').replace(' ', '')
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        v = float(s)
        return v
    except ValueError:
        return None


def parse_txt(path):
    with open(path, 'r', encoding='cp1252') as f:
        lines = f.read().splitlines()

    rows = []
    for line in lines:
        if not line.strip() or len(line) < 40:
            continue

        codigo = line[0:41].strip()
        desc = line[41:82].strip()
        cat = re.sub(r'^\d+-', '', line[82:123]).strip()
        marca = line[123:164].strip()
        prov = line[164:205].strip()

        existencia = None
        precio = None
        if len(line) > 205:
            m = TAIL_RE.match(line[205:])
            if m:
                existencia = parse_number(m.group(1))
                precio = parse_number(m.group(2))

        rows.append({
            'codigo': codigo,
            'desc': desc,
            'cat': cat,
            'marca': marca,
            'prov': prov,
            'existencia': existencia,
            'precio': precio,
        })
    return rows


def _localname(tag):
    return tag.rsplit('}', 1)[-1]


def _col_num(ref):
    letters = ''.join(ch for ch in ref if ch.isalpha())
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def _xml_bytes(root):
    return (b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            + ET.tostring(root, encoding='UTF-8'))


def _prepare_shared_strings(ss_data, reemplazos, valores):
    """Devuelve (bytes_sharedStrings, diccionario valor->indice).
    reemplazos = dict {texto_actual: texto_nuevo} aplicado a cadenas existentes."""
    sst = ET.fromstring(ss_data)
    sis = [e for e in sst if _localname(e.tag) == 'si']

    # Aplica reemplazos a las cadenas existentes (fecha, mensaje COMPRAS, etc.)
    for i, si in enumerate(sis):
        t_el = si.find(f'{{{NS}}}t')
        if t_el is not None and t_el.text in reemplazos:
            t_el.text = reemplazos[t_el.text]

    indice_por_valor = {}
    for i, si in enumerate(sis):
        t = si.findtext(f'{{{NS}}}t') or ''
        indice_por_valor[t] = i

    for valor in valores:
        if valor not in indice_por_valor:
            si = ET.SubElement(sst, f'{{{NS}}}si')
            ET.SubElement(si, f'{{{NS}}}t').text = valor
            indice_por_valor[valor] = len(sis)
            sis.append(si)

    total = len(sis)
    sst.set('count', str(total))
    sst.set('uniqueCount', str(total))

    return _xml_bytes(sst), indice_por_valor


def _fill_sheet(sheet_data, filas, string_idx, specs, borde_mapa=None):
    """Rellena las celdas de la hoja.
    filas = lista de dicts por producto.
    specs = lista de (num_columna, clave, tipo) con tipo 's' (texto) o 'n' (numero).
    borde_mapa = dict {num_columna: estilo_con_borde}. Si se pasa, las celdas
                 con valor quedan con borde y las vacias se eliminan (sin borde)."""
    root = ET.fromstring(sheet_data)
    sheet_data_el = root.find(f'{{{NS}}}sheetData')

    filas_por_numero = {}
    for row_el in sheet_data_el:
        if _localname(row_el.tag) == 'row':
            filas_por_numero[int(row_el.get('r'))] = row_el

    def obtener_fila(num):
        row_el = filas_por_numero.get(num)
        if row_el is None:
            row_el = ET.SubElement(sheet_data_el, f'{{{NS}}}row')
            row_el.set('r', str(num))
            filas_por_numero[num] = row_el
            return row_el
        return row_el

    def obtener_celda(row_el, ref, col_num):
        for child in row_el:
            if _localname(child.tag) == 'c' and child.get('r') == ref:
                return child
        # inserta en orden de columna
        cell = ET.Element(f'{{{NS}}}c')
        cell.set('r', ref)
        pos = 0
        for idx, child in enumerate(row_el):
            if _localname(child.tag) == 'c' and _col_num(child.get('r')) > col_num:
                pos = idx
                break
            pos = idx + 1
        row_el.insert(pos, cell)
        return cell

    def poner_valor(cell, valor, es_texto):
        if es_texto:
            cell.set('t', 's')
            v = cell.find(f'{{{NS}}}v')
            if v is None:
                v = ET.SubElement(cell, f'{{{NS}}}v')
            v.text = str(string_idx[valor])
        else:
            if 't' in cell.attrib:
                del cell.attrib['t']
            v = cell.find(f'{{{NS}}}v')
            if v is None:
                v = ET.SubElement(cell, f'{{{NS}}}v')
            v.text = repr(valor)

    for i, r in enumerate(filas):
        num = 13 + i
        row_el = obtener_fila(num)
        for col_num, clave, tipo in specs:
            valor = r.get(clave)
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                continue
            letra = chr(ord('A') + col_num - 1)
            poner_valor(obtener_celda(row_el, f'{letra}{num}', col_num), valor, tipo == 's')

    if borde_mapa is not None:
        variante, fallback = borde_mapa
        for row_el in filas_por_numero.values():
            if int(row_el.get('r')) < 13:
                continue
            # columna -> (celda, tiene_valor)
            celdas = {}
            for cell in list(row_el):
                if _localname(cell.tag) != 'c':
                    continue
                col = _col_num(cell.get('r'))
                if 2 <= col <= 7:
                    v = cell.find(f'{{{NS}}}v')
                    celdas[col] = (cell, v is not None and (v.text or '').strip() != '')
            # marco a la derecha: F con dato -> G; C con dato -> D
            conservar = set(col for col, (cell, tiene) in celdas.items() if tiene)
            if 6 in celdas and celdas[6][1]:
                conservar.add(7)
            if 3 in celdas and celdas[3][1]:
                conservar.add(4)
            for col, (cell, tiene) in celdas.items():
                s_attr = cell.get('s')
                if tiene:
                    if s_attr is None:
                        if col in fallback:
                            cell.set('s', str(fallback[col]))
                    else:
                        s_idx = int(s_attr)
                        if s_idx in variante:
                            cell.set('s', str(variante[s_idx]))
                else:
                    if col not in conservar:
                        row_el.remove(cell)
                    elif s_attr is None and col in fallback:
                        cell.set('s', str(fallback[col]))

    return _xml_bytes(root)


def _agregar_estilo(cellxfs, base_xf, border_id):
    """Copia un xf con un borderId nuevo, lo agrega a cellXfs y devuelve su indice."""
    nuevo = copy.deepcopy(base_xf)
    nuevo.set('borderId', str(border_id))
    nuevo.set('applyBorder', '1')
    cellxfs.append(nuevo)
    total = sum(1 for _ in cellxfs)
    cellxfs.set('count', str(total))
    return total - 1


def _estilos_borde(sheet_data, styles_data):
    """Devuelve (styles_bytes, (variante, fallback)) para bordes automaticos.
    variante: {estilo_original: estilo_con_borde} para estilos de celdas de datos
    (B..G) que no tienen borde.
    fallback: {columna: estilo_con_borde} para celdas sin estilo (filas fuera del
    rango estilizado de la plantilla)."""
    root = ET.fromstring(sheet_data)
    estilos_por_col = {}
    primer_estilo = {}
    for row_el in root.iter(f'{{{NS}}}row'):
        if int(row_el.get('r')) < 14:
            continue
        for cell in row_el:
            if _localname(cell.tag) != 'c':
                continue
            col = _col_num(cell.get('r'))
            if 2 <= col <= 7:
                s = int(cell.get('s', '0'))
                estilos_por_col.setdefault(col, set()).add(s)
                primer_estilo.setdefault(col, s)

    sroot = ET.fromstring(styles_data)
    cellxfs = None
    for el in sroot:
        if _localname(el.tag) == 'cellXfs':
            cellxfs = el
            break
    xfs = [xf for xf in cellxfs]

    variante = {}
    for col in range(2, 8):
        for s in estilos_por_col.get(col, ()):
            if int(xfs[s].get('borderId', '0')) == 0 and s not in variante:
                variante[s] = _agregar_estilo(cellxfs, xfs[s], 1)

    fallback = {}
    for col in range(2, 8):
        base = primer_estilo.get(col)
        if base is None:
            continue
        if int(xfs[base].get('borderId', '0')) == 0:
            fallback[col] = variante.get(base) or _agregar_estilo(cellxfs, xfs[base], 1)
        else:
            fallback[col] = base

    return _xml_bytes(sroot), (variante, fallback)


def _formato_bs(styles_data):
    """Cambia el formato de número de celdas de precio para usar formato venezolano:
    #.##0,00 (coma decimal, punto de miles) en lugar de #,##0.00 (formato US).
    Agrega un numFmt personalizado (numFmtId=165) y actualiza los estilos
    que usan numFmtId=4 (built-in #,##0.00) para usar el nuevo formato."""
    root = ET.fromstring(styles_data)

    numfmts = root.find(f'{{{NS}}}numFmts')
    if numfmts is None:
        numfmts = ET.SubElement(root, f'{{{NS}}}numFmts')
        root.insert(0, numfmts)

    ya_existe = any(
        nf.get('numFmtId') == '165'
        for nf in numfmts
        if _localname(nf.tag) == 'numFmt'
    )
    if not ya_existe:
        nf = ET.SubElement(numfmts, f'{{{NS}}}numFmt')
        nf.set('numFmtId', '165')
        nf.set('formatCode', '#,##0.00')
        numfmts.set('count', str(sum(1 for _ in numfmts)))

    cellxfs = None
    for el in root:
        if _localname(el.tag) == 'cellXfs':
            cellxfs = el
            break

    if cellxfs is not None:
        for xf in cellxfs:
            if _localname(xf.tag) != 'xf':
                continue
            if xf.get('numFmtId') == '4':
                xf.set('numFmtId', '165')

    return _xml_bytes(root)


def build_from_base(rows, base_path, out_path, reemplazos, specs, autoborder=False,
                    formato_bs=False):
    """Rellena la plantilla base con los productos y guarda out_path.
    autoborder=True deja los bordes solo en las celdas que tienen datos.
    formato_bs=True aplica formato #,##0.00 a los numeros."""
    valores = []
    for r in rows:
        for _, clave, tipo in specs:
            if tipo != 's':
                continue
            v = r.get(clave)
            if v and v not in valores:
                valores.append(v)

    z = zipfile.ZipFile(base_path)
    sheet_data = z.read('xl/worksheets/sheet1.xml')
    ss_bytes, string_idx = _prepare_shared_strings(
        z.read('xl/sharedStrings.xml'), reemplazos, valores)
    styles_data = z.read('xl/styles.xml')
    borde_mapa = None
    if autoborder:
        styles_data, borde_mapa = _estilos_borde(sheet_data, styles_data)
    if formato_bs:
        styles_data = _formato_bs(styles_data)
    sheet_bytes = _fill_sheet(sheet_data, rows, string_idx, specs, borde_mapa)

    try:
        if os.path.exists(out_path):
            os.remove(out_path)
    except PermissionError:
        pass

    try:
        with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as z2:
            for item in z.infolist():
                data = z.read(item.filename)
                if item.filename == 'xl/sharedStrings.xml':
                    data = ss_bytes
                elif item.filename == 'xl/worksheets/sheet1.xml':
                    data = sheet_bytes
                elif item.filename == 'xl/styles.xml':
                    data = styles_data
                z2.writestr(item, data)
    finally:
        z.close()

    return out_path


def ask_file_dialog():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='Selecciona el archivo TXT de lista de precios',
        filetypes=[('Archivos TXT', '*.txt'), ('Todos los archivos', '*.*')],
        initialdir=r'\\PRINCIPAL\a2admin\Empre001\REPORTS',
    )
    root.destroy()
    return path


def fmt_monto(v):
    """Formatea un monto con separador de miles español: 17325.0 -> '17.325'."""
    if v == int(v):
        s = f"{int(v):,}"
    else:
        s = f"{v:,.2f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def ask_excel_dialog():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='Selecciona el Excel de lista de precios',
        filetypes=[('Archivos Excel', '*.xlsx'), ('Todos los archivos', '*.*')],
        initialdir=CARPETA_SALIDAS if os.path.isdir(CARPETA_SALIDAS) else CARPETA_PROYECTO,
    )
    root.destroy()
    return path


def read_excel_rows(path):
    """Lee un Excel de lista (como el de la opcion 1) y devuelve filas con
    claves codigo, desc, modelo, marca, precio, cant."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    hdr_row = None
    cols = {}
    for r in range(1, 21):
        encontradas = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                u = v.strip().upper()
                u = u.replace('Í', 'I').replace('Ó', 'O').replace('É', 'E')
                u = u.replace('Á', 'A').replace('Ú', 'U').replace('Ñ', 'N')
                key = None
                if 'CODIGO' in u:
                    key = 'codigo'
                elif 'DESCRIPC' in u:
                    key = 'desc'
                elif u == 'MODELO':
                    key = 'modelo'
                elif u == 'MARCA':
                    key = 'marca'
                elif 'PRECIO' in u:
                    key = 'precio'
                elif u == 'CANT':
                    key = 'cant'
                if key:
                    encontradas[key] = c
        if encontradas:
            hdr_row = r
            cols = encontradas
            break

    if not hdr_row or 'codigo' not in cols:
        raise ValueError('No se encontro la tabla (Codigo/Descripcion/Precio) en el Excel.')

    filas = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        cod = ws.cell(r, cols['codigo']).value
        if cod is None or str(cod).strip() == '':
            continue
        precio = ws.cell(r, cols.get('precio', 6)).value
        if isinstance(precio, str):
            precio = parse_number(precio)
        cant = ws.cell(r, cols.get('cant', 7)).value
        if isinstance(cant, str):
            cant = parse_number(cant)
        filas.append({
            'codigo': str(cod).strip(),
            'desc': ws.cell(r, cols.get('desc', 3)).value,
            'modelo': ws.cell(r, cols.get('modelo', 4)).value,
            'marca': ws.cell(r, cols.get('marca', 5)).value,
            'precio': precio,
            'cant': cant,
        })
    return filas


def main_lista_bs():
    consola.titulo('OPCION 2 - LISTA EN Bs (EXCEL -> EXCEL)')

    base_path = os.path.join(CARPETA_PLANTILLAS, 'LISTA Base Bs.xlsx')

    if not os.path.exists(base_path):
        print(f'  No se encontro la plantilla: {base_path}')
        return

    origen = sys.argv[2] if len(sys.argv) >= 3 else None
    if not origen:
        origen = ask_excel_dialog()
        if not origen:
            print('  No seleccionaste ningun archivo. Cancelado.')
            return

    if len(sys.argv) >= 4:
        z = parse_monto(sys.argv[3])
    else:
        z = parse_monto(input('  Monto del dia de hoy? '))

    if z is None or z <= 0:
        print('  Monto invalido. Cancelado.')
        return

    try:
        rows = read_excel_rows(origen)
    except Exception as e:
        print(f'  Error al leer el Excel: {e}')
        return

    if not rows:
        print('  El Excel no tiene productos. Cancelado.')
        return

    for r in rows:
        if r['precio'] is not None:
            r['precio'] = round(r['precio'] * z, 2)

    fecha = date.today().strftime('%d/%m/%Y')
    x = 450 * z
    msg_viejo = 'COMPRAS A PARTIR DE:               X Bs'
    msg_nuevo = 'COMPRAS A PARTIR DE:               ' + fmt_monto(x) + ' Bs'

    reemplazos = {'DD/MM/AAAA': fecha, msg_viejo: msg_nuevo}
    specs = [
        (2, 'codigo', 's'),
        (3, 'desc', 's'),
        (4, 'modelo', 's'),
        (5, 'marca', 's'),
        (6, 'precio', 'n'),
        (7, 'cant', 'n'),
    ]

    nombre = 'LISTA ZM BS ' + date.today().strftime('%d-%m-%Y') + '.xlsx'
    os.makedirs(CARPETA_SALIDAS, exist_ok=True)
    out = os.path.join(CARPETA_SALIDAS, nombre)
    out = build_from_base(rows, base_path, out, reemplazos, specs, autoborder=True,
                          formato_bs=True)

    print('-' * 60)
    print(consola.verde(f'  OK: {len(rows)} productos x {z} -> {out}'))
    print(consola.verde(f'  Mensaje: {msg_nuevo}'))
    pdf = excel_a_pdf(out)
    if pdf:
        print(consola.verde(f'  PDF generado: {pdf}'))
    print()
    fecha_msg = date.today().strftime('%d-%m-%Y')
    print(consola.cian(consola.negrita(f'  LISTA DE PRECIOS Bs {fecha_msg}')))
    print(consola.cian(consola.negrita('  Los precios estan sujetos a cambio sin previo aviso')))
    print()
    consola.separador()

    archivos = [out]
    if pdf:
        archivos.append(pdf)
    _subir_a_drive(archivos, carpeta='folder_opcion2')


def _subir_a_drive(archivos, carpeta='folder_opcion1'):
    try:
        from GoogleDrive import subir_listas, _cargar_config
        config = _cargar_config()
        folder_id = config.get(carpeta)
        if not folder_id:
            print(consola.amarillo('  (No hay carpeta configurada en Google Drive)'))
            return
        print()
        print('  Subiendo archivos a Google Drive...')
        resultado = subir_listas(archivos, folder_id)
        if resultado is not None:
            print(consola.verde(f'  {resultado} archivos subidos a Google Drive'))
    except Exception as e:
        print(consola.amarillo(f'  (No se pudo subir a Google Drive: {e})'))


def _main_lista():
    args = sys.argv[1:]

    consola.titulo('OPCION 1 - LISTA DE PRECIOS (TXT -> EXCEL)')

    archivos_generados = []
    origen = args[0] if len(args) >= 1 else None

    base_path = os.path.join(CARPETA_PLANTILLAS, 'LISTA base.xlsx')

    if not os.path.exists(base_path):
        print(f'  No se encontro la plantilla: {base_path}')
        return

    if not origen:
        origen = ask_file_dialog()
        if not origen:
            print('  No seleccionaste ningun archivo. Cancelado.')
            return

    if os.path.isdir(origen):
        items = sorted(
            os.path.join(origen, f)
            for f in os.listdir(origen)
            if f.lower().endswith('.txt')
        )
    else:
        items = [origen]

    nombre = 'LISTA ZM ' + date.today().strftime('%d-%m-%Y') + '.xlsx'
    os.makedirs(CARPETA_SALIDAS, exist_ok=True)

    reemplazos = {'DD/MM/AAAA': date.today().strftime('%d/%m/%Y')}
    specs = [
        (2, 'codigo', 's'),
        (3, 'desc', 's'),
        (4, 'marca', 's'),
        (5, 'prov', 's'),
        (6, 'precio', 'n'),
    ]

    total = 0
    for idx, item in enumerate(items):
        rows = parse_txt(item)
        total_parse = len(rows)

        rows = [r for r in rows if r['existencia'] is None or r['existencia'] != 0]
        removed = total_parse - len(rows)

        rows.sort(key=lambda r: (r['cat'].lower(), r['codigo'].lower()))

        if len(items) > 1:
            base_nombre, ext = os.path.splitext(nombre)
            out = os.path.join(CARPETA_SALIDAS, f"{base_nombre} ({idx + 1}){ext}")
        else:
            out = os.path.join(CARPETA_SALIDAS, nombre)

        out = build_from_base(rows, base_path, out, reemplazos, specs,
                              autoborder=True, formato_bs=True)
        print(consola.verde(f'  OK: {total_parse} lineas -> {len(rows)} productos '
                            f'(se quitaron {removed} con existencia 0) -> {out}'))
        pdf = excel_a_pdf(out)
        if pdf:
            print(consola.verde(f'  PDF generado: {pdf}'))
        archivos_generados.append(out)
        if pdf:
            archivos_generados.append(pdf)
        fila_inicio = 13
        fila_fin = fila_inicio + len(rows) - 1
        print()
        print('=' * 60)
        print(consola.naranja(f'  Filas de items: desde la fila {fila_inicio} '
                              f'hasta la fila {fila_fin} ({len(rows)} filas)'))
        print('=' * 60)
        print()
        fecha_msg = date.today().strftime('%d-%m-%Y')
        print(consola.cian(consola.negrita(f'  Feliz tarde. Anexo LISTA ZM {fecha_msg}')))
        print(consola.cian(consola.negrita('  ACTUALIZACION DE INVENTARIO')))
        print()
        linea_log = (f'[{datetime.now().strftime("%d/%m/%Y %H:%M:%S")}] '
                     f'Filas de items: desde la fila {fila_inicio} '
                     f'hasta la fila {fila_fin} ({len(rows)} filas)')
        with open(LOG_FILAS, 'a', encoding='utf-8') as f:
            f.write(linea_log + '\n')
        total += len(rows)

    print('-' * 60)
    print(consola.cian(consola.negrita(f'  Total productos convertidos: {total}')))
    consola.separador()

    _subir_a_drive(archivos_generados)


def main():
    args = sys.argv[1:]

    if args and args[0] == '--bs':
        main_lista_bs()
        return

    _main_lista()


if __name__ == '__main__':
    main()
    try:
        input('Presiona Enter para cerrar...')
    except EOFError:
        pass
