# -*- coding: utf-8 -*-
"""
AnalizarCobranza.py
OPCION 4 - Analizar archivo de cobranza (PDF de CUENTAS POR COBRAR de ZOOM):

Para cada transaccion de cada cliente muestra:
    - Nombre del cliente que debe
    - Fecha de vencimiento
    - Dias de credito:  < 0 vigente, 0 vence hoy, > 0 vencido
    - No. de documento (la Nota)
    - Descripcion: si es solo un codigo (ej. "Doc : 00016671") avisa que la
      fecha de vencimiento posiblemente no este correcta (hay que actualizar
      la descripcion); si tiene un mensaje distinto, esta actualizada
    - Monto que debe
    - Estado y alerta

Genera informe en consola, archivo Markdown y archivo Excel.
Ademas, lista los clientes con dias >= -10.

Uso:
    python AnalizarCobranza.py                          (abre explorador)
    python AnalizarCobranza.py "C:/ruta/cobranza.pdf"
"""
import os
import re
import sys
from datetime import date

import pdfplumber

import consola

ENCABEZADOS = ('Código', 'Operación', 'CUENTAS', 'Vendedores', 'Total', 'Telefono')

CARPETA_CODIGO = os.path.dirname(os.path.abspath(__file__))
CARPETA_PROYECTO = os.path.dirname(CARPETA_CODIGO)
CARPETA_PDFS = os.path.join(CARPETA_PROYECTO, 'pdfs')
CARPETA_SALIDAS = os.path.join(CARPETA_PROYECTO, 'salidas')


def preguntar(texto):
    try:
        return input(texto)
    except EOFError:
        return ''


def ask_pdf_dialog():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='Selecciona el PDF de cobranza (CUENTAS POR COBRAR)',
        filetypes=[('Archivos PDF', '*.pdf'), ('Todos los archivos', '*.*')],
        initialdir=CARPETA_PDFS if os.path.isdir(CARPETA_PDFS) else CARPETA_PROYECTO,
    )
    root.destroy()
    return path


def _agrupar_lineas(words):
    words = sorted(words, key=lambda w: (w['top'], w['x0']))
    lineas = []
    for w in words:
        if lineas and abs(w['top'] - lineas[-1]['top']) < 3:
            lineas[-1]['words'].append(w)
        else:
            lineas.append({'top': w['top'], 'words': [w]})
    for ln in lineas:
        ln['words'].sort(key=lambda w: w['x0'])
    return lineas


def _es_transaccion(ws):
    if not ws:
        return False
    if ws[0]['x0'] >= 40:
        return False
    fechas = [w for w in ws if re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
    if len(fechas) < 2:
        return False
    xs = sorted(w['x0'] for w in fechas)
    return (45 <= xs[0] < 90) and (90 <= xs[1] < 150)


def _es_cliente1(ws):
    if not ws:
        return False
    if ws[0]['x0'] >= 40:
        return False
    if ws[0]['text'] in ENCABEZADOS:
        return False
    if any(re.match(r'^\d{2}/\d{2}/\d{4}$', w['text']) for w in ws):
        return False
    return any(100 <= w['x0'] < 300 for w in ws)


def _es_cliente2(ws):
    if not ws:
        return False
    if not all(w['x0'] >= 40 for w in ws):
        return False
    for w in ws:
        t = w['text']
        if '@' in t:
            return True
        if re.match(r'^[VEJPG]\s*-?\s*\d{5,}', t, re.IGNORECASE):
            return True
        if re.fullmatch(r'\d{6,}-\d', t):
            return True
        if re.fullmatch(r'[0-9][0-9\s\-()]*', t) and re.search(r'\d{3}', t):
            return True
    return False


def extraer_transacciones(path):
    """Devuelve lista de clientes, cada uno con sus transacciones."""
    clientes = []
    actual = None

    with pdfplumber.open(path) as pdf:
        for pag in pdf.pages:
            for ln in _agrupar_lineas(pag.extract_words()):
                ws = ln['words']
                texto = ' '.join(w['text'] for w in ws)
                if len(ws) == 1 and re.fullmatch(r'[\d.,]+', ws[0]['text']):
                    continue

                if _es_transaccion(ws):
                    if actual is None:
                        continue
                    fechas = [w for w in ws if re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
                    fechas.sort(key=lambda w: w['x0'])
                    emi = fechas[0]['text']
                    ven = fechas[1]['text']
                    dias_txt = next((w['text'] for w in ws if 150 <= w['x0'] < 190), '0')
                    doc = next((w['text'] for w in ws if 190 <= w['x0'] < 250), '')
                    monto_txt = next((w['text'] for w in ws if w['x0'] >= 500), '')
                    desc = ' '.join(w['text'] for w in ws if 250 <= w['x0'] < 500)
                    try:
                        monto = float(monto_txt.replace('.', '').replace(',', '.'))
                    except ValueError:
                        monto = None
                    try:
                        dias = int(dias_txt)
                    except ValueError:
                        dias = None
                    actual['transacciones'].append({
                        'emision': emi,
                        'vencimiento': ven,
                        'dias': dias,
                        'doc': doc,
                        'descripcion': desc.strip(),
                        'monto': monto,
                    })
                elif _es_cliente1(ws):
                    rif = next((w['text'] for w in ws if w['x0'] < 100), '')
                    nombre = ' '.join(w['text'] for w in ws if 100 <= w['x0'] < 290)
                    direccion = ' '.join(w['text'] for w in ws if w['x0'] >= 356)
                    activo = next((w['text'] for w in ws if 290 <= w['x0'] < 356), '')
                    actual = {
                        'nombre': nombre.strip(),
                        'rif': rif.strip(),
                        'direccion': direccion.strip(),
                        'activo': activo.strip(),
                        'telefono': '',
                        'email': '',
                        'transacciones': [],
                    }
                    clientes.append(actual)
                elif _es_cliente2(ws) and actual is not None:
                    telefono = ' '.join(w['text'] for w in ws if 100 <= w['x0'] < 300)
                    rif = next((w['text'] for w in ws if 300 <= w['x0'] < 390), '')
                    email = ' '.join(w['text'] for w in ws if w['x0'] >= 390)
                    if telefono:
                        actual['telefono'] = telefono.strip()
                    if rif and rif != telefono:
                        actual['rif'] = rif.strip()
                    if email:
                        actual['email'] = email.strip()
                # resto (encabezados, totales, pagina) se ignora

    for c in clientes:
        for t in c['transacciones']:
            t['nombre'] = c['nombre']
            t['rif'] = c['rif']
    return clientes


def descripcion_actualizada(desc):
    s = (desc or '').strip()
    if not s:
        return False
    if re.fullmatch(r'[Dd]oc\s*[:.]?\s*\d+', s):
        return False
    if re.fullmatch(r'0*\d{3,}', s):
        return False
    return True


def estado_transaccion(dias):
    if dias is None:
        return 'DESCONOCIDO', ''
    if dias > 0:
        return 'VENCIDO', f'VENCIDO hace {dias} dias'
    if dias == 0:
        return 'VENCE HOY', 'VENCE HOY'
    return 'VIGENTE', f'Vigente, restan {-dias} dias'


def alerta_transaccion(t):
    avisos = []
    if not descripcion_actualizada(t['descripcion']):
        avisos.append('DESCRIPCION NO ACTUALIZADA - verificar fecha de vencimiento')
    estado, _ = estado_transaccion(t['dias'])
    if estado == 'VENCIDO':
        avisos.append(f"VENCIDO ({t['dias']} dias)")
    elif estado == 'VENCE HOY':
        avisos.append('VENCE HOY')
    return ' | '.join(avisos)


def fmt_monto(v):
    if v is None:
        return '-'
    if v == int(v):
        s = f"{int(v):,}"
    else:
        s = f"{v:,.2f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def mensaje_por_vencer(t):
    """Mensaje para clientes por vencer (o que vencen hoy)."""
    monto = fmt_monto(t['monto'])
    if t['dias'] == 0:
        vence_linea = 'Se le vence HOY'
        se_vence = f'*Se vence el: {t["vencimiento"]} (HOY)*'
    else:
        vence_linea = f'Se le vence dentro de {-t["dias"]} días'
        se_vence = f'*Se vence el: {t["vencimiento"]}*'
    return '\n'.join([
        f'Al cliente: {t["nombre"]}',
        vence_linea,
        'Recuérdale el pago para que pueda disfrutar del descuento por pronto pago',
        '',
        'Esta debiendo',
        f'Nota: {t["doc"]} Monto: {monto} $',
        'Pendiente por aplicar $',
        '(5% de pronto pago)',
        se_vence,
    ])


def mensaje_vencido(t):
    """Mensaje para clientes con deuda vencida."""
    return '\n'.join([
        f'Solicitud de pago de: *{t["nombre"]}*',
        f'Nota: {t["doc"]} Monto: {fmt_monto(t["monto"])} $',
        f'*VENCIDA* desde el {t["vencimiento"]}',
        f'Tiene {t["dias"]} días vencida.',
        '¿Qué te ha dicho sobre el pago?',
    ])


def mensaje_cliente(t):
    """Devuelve el mensaje a enviar segun el estado de la transaccion."""
    if t['dias'] is None:
        return None
    if t['dias'] > 0:
        return mensaje_vencido(t)
    return mensaje_por_vencer(t)


def generar_informes(clientes, pdf_path):
    fecha_hoy = date.today()
    base_nombre = 'Informe Cobranza ' + fecha_hoy.strftime('%d-%m-%Y')
    os.makedirs(CARPETA_SALIDAS, exist_ok=True)
    ruta_md = os.path.join(CARPETA_SALIDAS, base_nombre + '.md')
    ruta_xlsx = os.path.join(CARPETA_SALIDAS, base_nombre + '.xlsx')

    total_trans = sum(len(c['transacciones']) for c in clientes)
    vencidos = sum(1 for c in clientes for t in c['transacciones']
                   if t['dias'] is not None and t['dias'] > 0)
    desc_no_act = sum(1 for c in clientes for t in c['transacciones']
                      if not descripcion_actualizada(t['descripcion']))
    total_monto = sum(t['monto'] for c in clientes for t in c['transacciones']
                      if t['monto'] is not None)

    _escribir_markdown(ruta_md, clientes, pdf_path, fecha_hoy,
                       total_trans, vencidos, desc_no_act, total_monto)
    _escribir_excel(ruta_xlsx, clientes)
    return ruta_md, ruta_xlsx


def _escribir_markdown(ruta_md, clientes, pdf_path, fecha_hoy,
                       total_trans, vencidos, desc_no_act, total_monto):
    lineas = []
    lineas.append(f'# Informe de Cobranza - {fecha_hoy.strftime("%d/%m/%Y")}')
    lineas.append('')
    lineas.append(f'- Fuente: `{os.path.basename(pdf_path)}`')
    lineas.append(f'- Clientes: **{len(clientes)}**')
    lineas.append(f'- Transacciones: **{total_trans}**')
    lineas.append(f'- Vencidos: **{vencidos}**')
    lineas.append(f'- Descripciones sin actualizar: **{desc_no_act}**')
    lineas.append(f'- Monto total: **{fmt_monto(total_monto)}**')
    lineas.append('')
    lineas.append('## Detalle')
    lineas.append('')

    for c in clientes:
        lineas.append(f'### {c["nombre"]}')
        lineas.append('')
        lineas.append(f'- RIF: {c["rif"] or "-"}  |  Telefono: {c["telefono"] or "-"}  |  '
                      f'Activo: {c["activo"] or "-"}')
        lineas.append('')
        lineas.append('| Vencimiento | Dias | Nota | Descripcion | Monto | Estado | Alerta |')
        lineas.append('|---|---|---|---|---|---|---|')
        for t in c['transacciones']:
            estado, _ = estado_transaccion(t['dias'])
            alerta = alerta_transaccion(t)
            if not descripcion_actualizada(t['descripcion']):
                desc_mostrar = f'{t["descripcion"]} ⚠'
            else:
                desc_mostrar = t['descripcion']
            lineas.append(
                f'| {t["vencimiento"]} | {t["dias"]} | {t["doc"]} | '
                f'{desc_mostrar} | {fmt_monto(t["monto"])} | {estado_md(estado)} | {alerta} |'
            )
        lineas.append('')

    lineas.append('## Listado de clientes con dias >= -10')
    lineas.append('')
    lineas.append('| Cliente | Vencimiento | Dias | Nota | Monto | Estado | Alerta |')
    lineas.append('|---|---|---|---|---|---|---|')
    filas = [(c, t) for c in clientes for t in c['transacciones']
             if t['dias'] is not None and t['dias'] >= -10]
    filas.sort(key=lambda ft: (ft[1]['dias'] or 0), reverse=True)
    for c, t in filas:
        estado, _ = estado_transaccion(t['dias'])
        lineas.append(
            f'| {c["nombre"]} | {t["vencimiento"]} | {t["dias"]} | {t["doc"]} | '
            f'{fmt_monto(t["monto"])} | {estado_md(estado)} | {alerta_transaccion(t)} |'
        )
    lineas.append('')

    filas_revisar = [(c, t) for c, t in filas
                     if not descripcion_actualizada(t['descripcion'])]
    lineas.append('## Descripcion sin actualizar (dias >= -10)')
    lineas.append('')
    lineas.append('Transacciones con la descripcion sin actualizar (solo un codigo): la')
    lineas.append('fecha de vencimiento posiblemente no este correcta. Revisalas antes de enviar.')
    lineas.append('')
    if filas_revisar:
        lineas.append('| Cliente | Emision | Vencimiento | Dias | Nota | Descripcion | '
                      'Monto | Estado | Alerta |')
        lineas.append('|---|---|---|---|---|---|---|---|---|')
        for c, t in filas_revisar:
            estado, _ = estado_transaccion(t['dias'])
            lineas.append(
                f'| {c["nombre"]} | {t["emision"]} | {t["vencimiento"]} | {t["dias"]} | '
                f'{t["doc"]} | {t["descripcion"]} | {fmt_monto(t["monto"])} | '
                f'{estado_md(estado)} | {alerta_transaccion(t)} |'
            )
    else:
        lineas.append('No hay transacciones pendientes de revisar.')
    lineas.append('')

    lineas.append('## Mensajes para clientes (dias >= -10)')
    lineas.append('')
    lineas.append('Bloques listos para copiar. Los marcados con **REVISAR** tienen la')
    lineas.append('descripcion sin actualizar: verifica la fecha de vencimiento antes de enviar.')
    lineas.append('')
    por_cliente = {}
    for c, t in filas:
        por_cliente.setdefault(c['nombre'], []).append((c, t))
    for nombre, items in por_cliente.items():
        lineas.append(f'### {nombre}')
        lineas.append('')
        for c, t in items:
            estado, _ = estado_transaccion(t['dias'])
            lineas.append(f'**{estado}** | Vence {t["vencimiento"]} | Dias {t["dias"]} | '
                          f'Nota {t["doc"]} | Monto {fmt_monto(t["monto"])}')
            lineas.append('')
            lineas.append('```')
            lineas.append(mensaje_cliente(t))
            lineas.append('```')
            if not descripcion_actualizada(t['descripcion']):
                lineas.append('')
                lineas.append('**⚠ REVISAR: descripcion no actualizada. '
                              'Verificar fecha de vencimiento antes de enviar.**')
            lineas.append('')

    with open(ruta_md, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lineas))
    return ruta_md


def _escribir_excel(ruta_xlsx, clientes):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    encabezados_detalle = ['Cliente', 'RIF', 'Vencimiento', 'Dias', 'Nota', 'Descripcion',
                           'Desc. Actualizada', 'Monto', 'Estado', 'Alerta']
    anchos_detalle = [38, 14, 12, 7, 11, 42, 15, 12, 11, 46]

    def hoja_detalle(ws, filas, encabezados, anchos):
        for col, h in enumerate(encabezados, 1):
            cel = ws.cell(1, col, h)
            cel.font = Font(bold=True, color='FFFFFF')
            cel.fill = PatternFill('solid', fgColor='1F4E78')
            cel.alignment = Alignment(horizontal='center')
        for r, fila in enumerate(filas, 2):
            for col, v in enumerate(fila, 1):
                cel = ws.cell(r, col, v)
                if isinstance(v, float):
                    cel.number_format = '#,##0.00'
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(encabezados))}{len(filas) + 1}'
        for i, a in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = a

    wb = Workbook()

    filas_todas = []
    for c in clientes:
        for t in c['transacciones']:
            estado, _ = estado_transaccion(t['dias'])
            filas_todas.append([
                c['nombre'], c['rif'], t['vencimiento'], t['dias'], t['doc'],
                t['descripcion'], 'SI' if descripcion_actualizada(t['descripcion']) else 'NO',
                t['monto'], estado, alerta_transaccion(t),
            ])
    ws1 = wb.active
    ws1.title = 'Detalle'
    hoja_detalle(ws1, filas_todas, encabezados_detalle, anchos_detalle)

    filas_filtro = [f for f in filas_todas
                    if isinstance(f[3], int) and f[3] >= -10]
    filas_filtro.sort(key=lambda f: f[3], reverse=True)
    ws2 = wb.create_sheet('>= -10')
    hoja_detalle(ws2, filas_filtro, encabezados_detalle, anchos_detalle)

    encabezados_revisar = ['Cliente', 'Emision', 'Vencimiento', 'Dias', 'Nota',
                           'Descripcion', 'Monto', 'Estado', 'Alerta']
    anchos_revisar = [38, 12, 12, 7, 11, 42, 12, 11, 46]
    filas_revisar = []
    for c in clientes:
        for t in c['transacciones']:
            if (t['dias'] is not None and t['dias'] >= -10
                    and not descripcion_actualizada(t['descripcion'])):
                estado, _ = estado_transaccion(t['dias'])
                filas_revisar.append([
                    c['nombre'], t['emision'], t['vencimiento'], t['dias'], t['doc'],
                    t['descripcion'], t['monto'], estado, alerta_transaccion(t),
                ])
    filas_revisar.sort(key=lambda f: f[3], reverse=True)
    ws3 = wb.create_sheet('Desc. sin act. (>= -10)')
    hoja_detalle(ws3, filas_revisar, encabezados_revisar, anchos_revisar)

    wb.save(ruta_xlsx)
    return ruta_xlsx


def _pintar_estado(estado):
    if estado == 'VENCIDO':
        return consola.rojo(estado)
    if estado == 'VENCE HOY':
        return consola.amarillo(estado)
    if estado == 'VIGENTE':
        return consola.verde(estado)
    return estado


def _pintar_alerta(alerta):
    if not alerta:
        return ''
    if 'VENCIDO' in alerta:
        return consola.rojo(alerta)
    if 'VENCE HOY' in alerta:
        return consola.amarillo(alerta)
    return consola.amarillo(alerta)


def estado_md(estado):
    """Estado en negrita para el Markdown cuando es importante."""
    if estado in ('VENCIDO', 'VENCE HOY'):
        return f'**{estado}**'
    return estado


def mostrar_consola(clientes, pdf_path):
    print('\n' + '=' * 78)
    print('  ' + consola.cian(consola.negrita('INFORME DE COBRANZA')))
    print(f'  Fuente: {os.path.basename(pdf_path)}')
    print('=' * 78)

    total_trans = sum(len(c['transacciones']) for c in clientes)
    vencidos = sum(1 for c in clientes for t in c['transacciones']
                   if t['dias'] is not None and t['dias'] > 0)
    desc_no_act = sum(1 for c in clientes for t in c['transacciones']
                      if not descripcion_actualizada(t['descripcion']))
    total_monto = sum(t['monto'] for c in clientes for t in c['transacciones']
                      if t['monto'] is not None)
    print(f'  Clientes: {consola.negrita(len(clientes))}  |  '
          f'Transacciones: {consola.negrita(total_trans)}  |  '
          f'Vencidos: {consola.rojo(vencidos)}  |  '
          f'Desc. sin actualizar: {consola.amarillo(desc_no_act)}  |  '
          f'Monto total: {consola.negrita(fmt_monto(total_monto))}')

    for c in clientes:
        print('\n' + '-' * 78)
        print('  ' + consola.cian(consola.negrita(f'CLIENTE: {c["nombre"]}')))
        print(f'  RIF: {c["rif"] or "-"}   Telefono: {c["telefono"] or "-"}   '
              f'Activo: {c["activo"] or "-"}')
        for t in c['transacciones']:
            estado, msg = estado_transaccion(t['dias'])
            alerta = alerta_transaccion(t)
            print(f'  Emision: {t["emision"]}  Vence: {t["vencimiento"]}  '
                  f'Dias: {t["dias"]}  Nota: {t["doc"]}  '
                  f'Monto: {consola.negrita(fmt_monto(t["monto"]))}  '
                  f'{_pintar_estado(estado)}')
            print(f'    Descripcion: {t["descripcion"]}')
            if not descripcion_actualizada(t['descripcion']):
                print('    ' + consola.amarillo('!! AVISO: descripcion es solo un codigo, '
                      'la fecha de vencimiento posiblemente no este correcta. '
                      'Actualizar descripcion.'))
            if alerta:
                print(f'    [{_pintar_alerta(alerta)}]')
            print()

    print('\n' + '=' * 78)
    print('  ' + consola.cian(consola.negrita('LISTADO DE CLIENTES CON DIAS >= -10')))
    print('=' * 78)
    filas = [(c, t) for c in clientes for t in c['transacciones']
             if t['dias'] is not None and t['dias'] >= -10]
    filas.sort(key=lambda ft: (ft[1]['dias'] or 0), reverse=True)
    for c, t in filas:
        estado, _ = estado_transaccion(t['dias'])
        monto_mostrar = consola.negrita(f'{fmt_monto(t["monto"]):>12}')
        print(f'  {c["nombre"][:38]:<38}  Vence {t["vencimiento"]}  '
              f'Dias {t["dias"]:>4}  Nota {t["doc"]}  '
              f'{monto_mostrar}  {_pintar_estado(estado)}')
    print()

    print('=' * 78)
    print('  ' + consola.amarillo(consola.negrita(
        'DESCRIPCION SIN ACTUALIZAR (DIAS >= -10)')))
    print('=' * 78)
    filas_revisar = [(c, t) for c, t in filas
                     if not descripcion_actualizada(t['descripcion'])]
    if not filas_revisar:
        print('  No hay transacciones pendientes de revisar.')
    else:
        for c, t in filas_revisar:
            estado, _ = estado_transaccion(t['dias'])
            print(f'  {c["nombre"][:34]:<34}  Emision {t["emision"]}  '
                  f'Vence {t["vencimiento"]}  Dias {t["dias"]:>4}  '
                  f'Nota {t["doc"]}  {fmt_monto(t["monto"]):>10}  '
                  f'{_pintar_estado(estado)}')
    print()


def main():
    args = sys.argv[1:]
    pdf_path = args[0] if args and args[0].lower().endswith('.pdf') else None
    if not pdf_path:
        pdf_path = ask_pdf_dialog()
    if not pdf_path or not os.path.exists(pdf_path):
        print('  No seleccionaste ningun PDF. Cancelado.')
        return

    print('Leyendo el PDF...')
    clientes = extraer_transacciones(pdf_path)
    if not clientes:
        print('  No se pudo leer la cobranza del PDF. Cancelado.')
        return

    mostrar_consola(clientes, pdf_path)
    ruta_md, ruta_xlsx = generar_informes(clientes, pdf_path)
    print('=' * 78)
    print(f'  Informe Markdown: {ruta_md}')
    print(f'  Informe Excel:    {ruta_xlsx}')
    print('=' * 78)
    print('  Mensajes para clientes (vencidos, vence hoy y por vencer):')
    print('  seccion "Mensajes para clientes" del archivo Markdown.')
    print('=' * 78)


if __name__ == '__main__':
    main()
    try:
        input('Presiona Enter para cerrar...')
    except EOFError:
        pass
