# -*- coding: utf-8 -*-
"""
CompararVentas.py
OPCION 8 - NO VENDIDOS: compara ventas contra un pedido y lo guarda en Supabase:
    Lee uno o varios PDFs de VENTAS (los vendidos), cada uno asociado a un CLIENTE
    (nombre detectado del PDF o confirmado por el usuario), y un PEDIDO (PDF o Excel).
    El pedido tambien puede agregarse a mano (A): nombre del producto (en MAYUSCULAS)
    y cantidad; al terminar se guarda y se sube a Supabase.

    Para cada cliente calcula:
        No vendido = Pedido - Vendido  (minimo 0)

    Los codigos del pedido que no aparecen en las ventas del cliente se listan
    como no vendidos completos. Muestra en el terminal el CODIGO y la CANTIDAD
    (sin descripcion).

    Guarda los resultados por MES -> CLIENTE -> (codigo, cantidad) en Supabase
    (tabla public.no_vendidos) y guarda un Excel por cliente en salidas.

OPCION 9 - consultar_historial():
    Pide un mes y muestra todos los clientes con los items que no se les vendio.

Uso:
    python CompararVentas.py                          (abre explorador)
    python CompararVentas.py "C:/ruta/ventas.pdf" "C:/ruta/pedido.xlsx"
"""
import os
import re
import sys
from datetime import date

import consola
from ConvertirListaExcel import parse_number, read_excel_rows
from supabase_client import (SupabaseError, upsert_no_vendido,
                             listar_no_vendidos_mes, listar_meses_no_vendidos)

CARPETA_CODIGO = os.path.dirname(os.path.abspath(__file__))
CARPETA_PROYECTO = os.path.dirname(CARPETA_CODIGO)
CARPETA_PDFS = os.path.join(CARPETA_PROYECTO, 'pdfs')
CARPETA_SALIDAS = os.path.join(CARPETA_PROYECTO, 'salidas')

MESES = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
         7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre',
         12: 'Diciembre'}


def _mes_nombre(mes):
    """'2026-04' -> 'Abril 2026'."""
    try:
        anio, num = str(mes).split('-')
        return f'{MESES.get(int(num), num)} {anio}'
    except Exception:
        return str(mes)


def ask_pdf_dialog(titulo):
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title=titulo,
        filetypes=[('Archivos PDF', '*.pdf'), ('Todos los archivos', '*.*')],
        initialdir=CARPETA_PDFS if os.path.isdir(CARPETA_PDFS) else CARPETA_PROYECTO,
    )
    root.destroy()
    return path


def ask_pedido_dialog():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='Selecciona el PEDIDO (PDF o Excel)',
        filetypes=[('PDF o Excel', '*.pdf;*.xlsx'), ('Archivos PDF', '*.pdf'),
                   ('Archivos Excel', '*.xlsx'), ('Todos los archivos', '*.*')],
        initialdir=CARPETA_PDFS if os.path.isdir(CARPETA_PDFS) else CARPETA_PROYECTO,
    )
    root.destroy()
    return path


def _agrupar_lineas(words):
    words = sorted(words, key=lambda w: (w['top'], w['x0']))
    lineas = []
    for w in words:
        if lineas and abs(w['top'] - lineas[-1]['top']) < 3:
            lineas[-1]['words'].append(w)
        else:
            lineas.append({'top': w['top'], 'words': [w]})
    for ln in lineas:
        ln['words'].sort(key=lambda w: w['x0'])
    return lineas


def _norm(t):
    return t.upper().replace('Ó', 'O').replace('Í', 'I').replace('Á', 'A') \
             .replace('\ufffd', '').strip()


PALABRAS_GENERICAS = ('VENTA', 'VENTAS', 'FACTURA', 'NOTA', 'REPORTE', 'LISTA',
                      'PEDIDO', 'PDF', 'CLIENTE', 'ZOOM', 'COPIA', 'PAGO',
                      'COBRANZA', 'PRECIO', 'ENVIADA', 'ENVIO')


def _limpiar_nombre_cliente(nombre):
    """Limpia un nombre de archivo o linea para quedarse solo con el cliente:
    quita extension, fechas, numeros, separadores y palabras genericas."""
    s = str(nombre)
    s = re.sub(r'\.(pdf|xlsx|xls|txt)$', '', s, flags=re.I)
    s = re.sub(r'\b\d{4}[-/]\d{2}[-/]\d{2}\b', ' ', s)
    s = re.sub(r'\b\d{2}[-/]\d{2}[-/]\d{2,4}\b', ' ', s)
    s = re.sub(r'\b\d{1,2}[-/]\d{1,2}\b', ' ', s)
    s = s.replace('_', ' ')
    s = re.sub(r'[.,;:()\[\]{}]', ' ', s)
    s = re.sub(r'\b\d+\b', ' ', s)
    s = s.replace('-', ' ')
    tokens = []
    for t in s.split():
        u = _norm(t)
        if len(t) <= 1 or u in PALABRAS_GENERICAS:
            continue
        # quita codigos: token con letras y numeros mezclados (N16756, F4663M10, K80028)
        if any(c.isalpha() for c in t) and any(c.isdigit() for c in t):
            continue
        tokens.append(t)
    return ' '.join(tokens).strip()


def _texto_pdf(path, lineas_max=12):
    """Devuelve las primeras lineas de texto de un PDF."""
    import pdfplumber

    lineas = []
    try:
        with pdfplumber.open(path) as pdf:
            for pag in pdf.pages[:3]:
                for ln in _agrupar_lineas(pag.extract_words()):
                    lineas.append(' '.join(w['text'] for w in ln['words']).strip())
                    if len(lineas) >= lineas_max:
                        return lineas
    except Exception:
        pass
    return lineas


def _sugerir_cliente_pdf(path):
    """Sugiere el cliente desde el NOMBRE DEL ARCHIVO (normalmente lo contiene)
    y, si no, desde las primeras lineas del PDF."""
    limpiado = _limpiar_nombre_cliente(os.path.basename(path))
    if limpiado:
        return limpiado

    excluir = ('CODIGO', 'DESCRIPCION', 'CANTIDAD', 'CANT', 'FECHA', 'REPORTE',
               'PAGINA', 'TOTAL', 'PRECIO', 'VENTA', 'LISTA', 'NOTA', 'FACTURA',
               'COMPROBANTE', 'CLIENTE', 'DIRECCION', 'RIF', 'TELEFONO')
    for t in _texto_pdf(path):
        u = _norm(t)
        if len(t) >= 3 and not any(k in u for k in excluir) and any(c.isalpha() for c in t):
            return t
    return ''


def extraer_items_pdf(path):
    """Extrae items (codigo, cant) de un PDF tipo tabla de lista.
    No toma la descripcion. Detecta el encabezado CODIGO / CANT.
    Si no lo encuentra usa una heuristica."""
    import pdfplumber

    items = []
    codigo_x = None
    cant_x = None
    con_encabezado = False

    with pdfplumber.open(path) as pdf:
        for pag in pdf.pages:
            for ln in _agrupar_lineas(pag.extract_words()):
                texto = ' '.join(w['text'] for w in ln['words'])
                u = _norm(texto)
                if codigo_x is None:
                    for w in ln['words']:
                        t = _norm(w['text'])
                        if t in ('CODIGO', 'CDIGO'):
                            codigo_x = w['x0']
                        elif t in ('CANT', 'CANTIDAD', 'UNID', 'UNIDADES', 'QTY'):
                            cant_x = w['x0']
                    if codigo_x is not None:
                        con_encabezado = True
                        continue
                if 'TOTAL' in u or 'PAGINA' in u:
                    continue

                if codigo_x is not None:
                    code_words = [w for w in ln['words'] if abs(w['x0'] - codigo_x) < 12]
                    codigo = ''.join(w['text'] for w in
                                     sorted(code_words, key=lambda w: w['x0'])).strip()
                    if not codigo or not any(ch.isalnum() for ch in codigo):
                        continue
                    if not any(ch.isdigit() for ch in codigo):
                        continue
                    cant = None
                    if cant_x is not None:
                        candidatos = [(abs(w['x0'] - cant_x), w) for w in ln['words']
                                      if parse_number(w['text']) is not None]
                        if candidatos:
                            _, w_min = min(candidatos, key=lambda t: t[0])
                            cant = parse_number(w_min['text'])
                    if cant is None:
                        nums = [parse_number(w['text']) for w in ln['words']
                                if parse_number(w['text']) is not None]
                        if nums:
                            cant = nums[-1]
                    if cant is not None:
                        items.append({'codigo': codigo, 'desc': '', 'cant': cant})
                else:
                    # Heuristica sin encabezado: primer token = codigo,
                    # ultimo numero = cantidad, sin descripcion
                    nums = [w for w in ln['words'] if parse_number(w['text']) is not None]
                    if len(nums) < 1 or not nums[-1]['text'].strip():
                        continue
                    cant = parse_number(nums[-1]['text'])
                    prim = ln['words'][0]
                    if prim is nums[-1] or not prim['text'][0].isalpha():
                        continue
                    codigo = prim['text'].strip()
                    if not any(ch.isdigit() for ch in codigo):
                        continue
                    items.append({'codigo': codigo, 'desc': '', 'cant': cant})
    return items, con_encabezado


def extraer_items_excel(path):
    """Lee un Excel tipo lista y devuelve items (codigo, cant). Sin descripcion."""
    filas = read_excel_rows(path)
    return [{'codigo': r['codigo'], 'desc': '', 'cant': _cant_num(r['cant'])}
            for r in filas]


def _cant_num(cant):
    if isinstance(cant, str):
        cant = parse_number(cant)
    return cant if cant is not None else 0


def cargar_items(path):
    """Devuelve (items, descripcion_tipo)."""
    ext = path.lower()
    if ext.endswith('.xlsx'):
        return extraer_items_excel(path), 'Excel'
    if ext.endswith('.pdf'):
        items, con = extraer_items_pdf(path)
        if not con:
            print(consola.amarillo('  Aviso: el PDF no tiene encabezado CODIGO/CANT; '
                                   'se uso el lector generico.'))
        return items, 'PDF'
    raise ValueError('Formato no soportado (use PDF o Excel .xlsx).')


def seleccionar_pdfs_ventas():
    """Pide uno o mas PDFs de ventas y el nombre del cliente de cada uno.
    Devuelve lista de (ruta, cliente)."""
    pdfs = []
    while True:
        n = len(pdfs) + 1
        p = ask_pdf_dialog(f'Selecciona el PDF de VENTAS #{n}')
        if not p:
            if pdfs:
                break
            return []
        sug = _sugerir_cliente_pdf(p)
        if sug:
            print(f'  Cliente detectado en el PDF: {consola.cian(sug)}')
        texto = (f'  Nombre del cliente de este PDF'
                 f'{" (Enter usa la sugerencia)" if sug else ""}: ')
        cliente = input(texto).strip()
        if not cliente and sug:
            cliente = sug
        if not cliente:
            cliente = f'Cliente {n}'
        pdfs.append((p, cliente))
        mas = input(f'  PDF {n} agregado (Cliente: {consola.cian(cliente)}). '
                    f'Agregar otro PDF de ventas? [s/N]: ').strip().lower()
        if mas not in ('s', 'si', 'y', 'yes'):
            break
    return pdfs


def _fmt_cant(n):
    """5.0 -> '5', 5.5 -> '5.5'."""
    try:
        if n == int(n):
            return str(int(n))
    except (TypeError, ValueError):
        pass
    return str(n)


def agregar_pedido_manual():
    """Permite introducir a mano el PEDIDO: NOMBRE DEL PRODUCTO y CANTIDAD.
    El nombre se guarda siempre en MAYUSCULAS.
    Acepta 'NOMBRE CANTIDAD' en una sola linea (ej: FARO DELANTERO 4)
    o nombre y cantidad por separado.
    Devuelve lista de items {'codigo': nombre, 'cant': cantidad}."""
    items = {}
    print()
    print('  PEDIDO MANUAL: escribe el NOMBRE del producto y la cantidad.')
    print('  En una sola linea:   FARO DELANTERO 4')
    print('  (No importa si es mayuscula o minuscula; se guarda en MAYUSCULAS)')
    while True:
        entrada = input('  Producto y cantidad (Enter para terminar): ').strip()
        if not entrada:
            break
        partes = entrada.split()
        cantidad = parse_number(partes[-1]) if partes else None
        if len(partes) >= 2 and cantidad is not None and cantidad > 0:
            nombre = ' '.join(partes[:-1]).upper()
        else:
            nombre = entrada.upper()
            cant = input('  Cantidad: ').strip()
            cantidad = parse_number(cant)
        if not nombre:
            continue
        if cantidad is None or cantidad <= 0:
            print('    Cantidad invalida. Intenta de nuevo.')
            continue
        it = items.setdefault(nombre, {'codigo': nombre, 'cant': 0})
        it['cant'] += cantidad
        print(f'    Agregado: {nombre} = {_fmt_cant(cantidad)}')
    return list(items.values())


def calcular_no_vendidos(sold, pedido):
    """sold = dict codigo -> cantidad (o {cant}). pedido = dict codigo -> {cant}.
    Devuelve lista de dicts {codigo, pedido, vendido, no_vendido}.
    No se usa la descripcion de los items."""
    resultado = []
    for codigo, item in pedido.items():
        ped = item['cant']
        ven = sold.get(codigo, 0)
        if isinstance(ven, dict):
            ven = ven.get('cant', 0)
        no_vendido = int(round(max(0, ped - ven)))
        resultado.append({
            'codigo': codigo,
            'pedido': ped,
            'vendido': ven,
            'no_vendido': no_vendido,
        })
    resultado.sort(key=lambda r: r['no_vendido'], reverse=True)
    return resultado


def guardar_excel(filas, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    encabezados = ['CODIGO', 'PEDIDO', 'VENDIDO', 'NO VENDIDO']
    anchos = [18, 10, 10, 12]

    wb = Workbook()
    ws = wb.active
    ws.title = 'No vendidos'
    for col, h in enumerate(encabezados, 1):
        cel = ws.cell(1, col, h)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = PatternFill('solid', fgColor='7B2D26')
        cel.alignment = Alignment(horizontal='center')
    for r, fila in enumerate(filas, 2):
        ws.cell(r, 1, fila['codigo'])
        ws.cell(r, 2, fila['pedido'])
        ws.cell(r, 3, fila['vendido'])
        ws.cell(r, 4, fila['no_vendido'])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(encabezados))}{len(filas) + 1}'
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    wb.save(out_path)
    return out_path


def _nombre_archivo(s):
    invalidos = '<>:"/\\|?*'
    return ''.join(c for c in str(s) if c not in invalidos).strip() or 'Cliente'


def guardar_cliente_supabase(mes, cliente, items):
    """Guarda en Supabase los no vendidos de un cliente en un mes.
    items = lista de dicts {codigo, no_vendido}. Devuelve cantidad guardada."""
    guardadas = 0
    for r in items:
        upsert_no_vendido({
            'mes': mes,
            'cliente': cliente,
            'codigo': r['codigo'],
            'cantidad': r['no_vendido'],
        })
        guardadas += 1
    return guardadas


def consultar_historial():
    """OPCION 9: muestra los no vendidos de un mes por cliente (Supabase)."""
    consola.titulo('OPCION 9 - NO VENDIDOS: HISTORIAL POR MES', ancho=58)

    try:
        meses = listar_meses_no_vendidos()
    except SupabaseError as e:
        print(consola.rojo(f'  Error de Supabase: {e}'))
        print('  Revisa config/supabase.json y que exista la tabla no_vendidos.')
        return

    if not meses:
        print('  No hay registros en Supabase todavia.')
        print('  Genera los no vendidos en la opcion 8 para guardarlos.')
        return

    print('  Meses con registros:')
    for m in meses:
        print(f'    - {m}  ({_mes_nombre(m)})')

    raw = input(f'  Mes a consultar [Enter = {meses[0]}]: ').strip() or meses[0]
    if raw not in meses:
        for m in meses:
            if _mes_nombre(m).lower() == raw.lower():
                raw = m
                break
    mes = raw
    if mes not in meses:
        print(f'  No hay registros del mes {mes}.')
        return

    try:
        filas = listar_no_vendidos_mes(mes)
    except SupabaseError as e:
        print(consola.rojo(f'  Error de Supabase: {e}'))
        return

    por_cliente = {}
    for f in filas:
        cliente = f.get('cliente') or '?'
        por_cliente.setdefault(cliente, []).append(
            (f.get('codigo') or '', int(round(f.get('cantidad') or 0))))

    if not por_cliente:
        print(f'  No hay registros del mes {mes}.')
        return

    print()
    print('  MES = ' + consola.cian(consola.negrita(_mes_nombre(mes))))
    print('=' * 60)
    for cliente in sorted(por_cliente, key=str.lower):
        items = por_cliente[cliente]
        print()
        print('-' * 60)
        print(consola.negrita(f'  {cliente}'))
        print('-' * 60)
        print()
        total = 0
        for codigo, cantidad in sorted(items, key=lambda x: x[1], reverse=True):
            total += cantidad
            print(f'    {codigo} = {consola.rojo(cantidad)}')
        print()
        print(consola.amarillo(f'  Total no vendido: {total}  |  Items: {len(items)}'))
        print('-' * 60)
    print()
    print('=' * 60)
    print(consola.cian(consola.negrita(
        f'  RESUMEN {_mes_nombre(mes)}: {len(por_cliente)} clientes en el registro.')))
    print('=' * 60)


def main():
    args = sys.argv[1:]

    consola.titulo('OPCION 8 - NO VENDIDOS (VENTAS vs PEDIDO)')

    if len(args) >= 2:
        cliente = input('  Nombre del cliente de este PDF de ventas: ').strip() or 'Cliente 1'
        pdfs = [(args[0], cliente)]
    else:
        pdfs = seleccionar_pdfs_ventas()
        if not pdfs:
            print('  No seleccionaste ningun PDF de ventas. Cancelado.')
            return

    for p, _ in pdfs:
        if not os.path.exists(p):
            print(f'  No se encontro el archivo: {p}')
            return

    sold_total = {}
    por_cliente = {}
    for path, cliente in pdfs:
        try:
            items, _ = cargar_items(path)
        except Exception as e:
            print(f'  Error al leer {path}: {e}')
            return
        ventas = por_cliente.setdefault(cliente, {})
        for it in items:
            cod = it['codigo']
            ventas[cod] = ventas.get(cod, 0) + it['cant']
            sold_total[cod] = sold_total.get(cod, 0) + it['cant']
        print(f'  VENTAS {os.path.basename(path)} (Cliente: {cliente}): '
              f'{len(items)} items leidos.')
        for it in sorted(items, key=lambda x: x['codigo']):
            print(f'      {it["codigo"]:<22} = {_fmt_cant(it["cant"])}')

    print()
    print('  PEDIDO (las cosas que se pidieron)')
    if len(args) >= 2:
        modo = 'b'
    else:
        modo = input('  Agregar pedido a mano [A] o por documento [B]? [A/B]: ').strip().lower()
    if modo in ('a', 'mano', 'manual'):
        ped_items = agregar_pedido_manual()
        tipo_ped = 'MANUAL'
        pedido_path = None
        if not ped_items:
            print('  El pedido manual esta vacio. Cancelado.')
            return
    else:
        if len(args) >= 2:
            pedido_path = args[1]
        else:
            pedido_path = ask_pedido_dialog()
            if not pedido_path:
                print('  No seleccionaste el pedido. Cancelado.')
                return
        if not os.path.exists(pedido_path):
            print(f'  No se encontro el archivo: {pedido_path}')
            return
        try:
            ped_items, tipo_ped = cargar_items(pedido_path)
        except Exception as e:
            print(f'  Error al leer el pedido: {e}')
            return
    if not ped_items:
        print('  El pedido no tiene items. Cancelado.')
        return
    pedido = {it['codigo']: it for it in ped_items}
    if pedido_path:
        print(f'  PEDIDO {os.path.basename(pedido_path)} ({tipo_ped}): '
              f'{len(ped_items)} items leidos.')
    else:
        print(f'  PEDIDO MANUAL ({tipo_ped}): {len(ped_items)} items.')

    resultado = calcular_no_vendidos(sold_total, pedido)
    no_vendidos = [r for r in resultado if r['no_vendido'] > 0]
    completos = sum(1 for r in resultado if r['no_vendido'] == 0)

    print()
    print('=' * 72)
    print(consola.rojo(consola.negrita('  ITEMS DEL PEDIDO QUE NO SE VENDIERON')))
    print('=' * 72)
    if not no_vendidos:
        print(consola.verde('  Todo el pedido se vendio. No hay items sin vender.'))
    else:
        for r in no_vendidos:
            print(f'  {r["codigo"]:<20} {consola.rojo(r["no_vendido"]):>8}')
    print('-' * 72)
    print(f'  Pedido total: {consola.negrita(len(resultado))} items  |  '
          f'Vendidos completos: {consola.verde(completos)}  |  '
          f'No vendidos: {consola.rojo(len(no_vendidos))}')
    print('=' * 72)

    mes = date.today().strftime('%Y-%m')
    os.makedirs(CARPETA_SALIDAS, exist_ok=True)

    print()
    print('=' * 72)
    print(consola.cian(consola.negrita('  GUARDADO EN SUPABASE (MES: %s)' % mes)))
    print('=' * 72)
    total_registros = 0
    for cliente in sorted(por_cliente, key=str.lower):
        ventas = por_cliente[cliente]
        res = calcular_no_vendidos(ventas, pedido)
        items_nv = [r for r in res if r['no_vendido'] > 0]
        try:
            guardadas = guardar_cliente_supabase(mes, cliente, items_nv)
            total_registros += guardadas
            print(consola.verde(f'  Cliente {cliente}: {guardadas} no vendidos guardados.'))
        except SupabaseError as e:
            print(consola.rojo(f'  Error Supabase ({cliente}): {e}'))

        if items_nv:
            nombre = ('No Vendidos ' + _nombre_archivo(cliente) + ' '
                      + date.today().strftime('%d-%m-%Y') + '.xlsx')
            out = os.path.join(CARPETA_SALIDAS, nombre)
            guardar_excel(items_nv, out)
            print(consola.verde(f'  OK: {cliente} -> {out}'))

    if not no_vendidos:
        nombre = 'No Vendidos ' + date.today().strftime('%d-%m-%Y') + '.xlsx'
        out = os.path.join(CARPETA_SALIDAS, nombre)
        guardar_excel(no_vendidos, out)
        print(consola.verde(f'  OK: guardado -> {out}'))

    if total_registros:
        print(consola.verde(f'  Total registros guardados en Supabase: {total_registros}'))
    print('=' * 72)
    consola.separador()


if __name__ == '__main__':
    main()
    try:
        input('Presiona Enter para cerrar...')
    except EOFError:
        pass
